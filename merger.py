import click
import os
from openpyxl import Workbook, load_workbook


@click.command()
@click.argument('files', nargs=-1, type=click.Path())
def merge(files):
    """ Merge xlsx files to one. """

    wb = Workbook()

    for filename in files:
        if os.path.isdir(filename):
            for filename_ in os.listdir(filename):
                if filename_.endswith('.xlsx'):
                    click.echo(filename + "\\" + filename_)
                    wb_load = load_workbook(filename + "\\" + filename_, keep_vba=True)
                    for ws in wb_load.worksheets:
                        ws2 = wb.create_sheet()
                        for row in ws:
                            for cell in row:
                                ws2[cell.coordinate].value = cell.value
        else:
            click.echo(filename)
            wb_load = load_workbook(filename, keep_vba=True)
            for ws in wb_load.worksheets:
                ws2 = wb.create_sheet()
                for row in ws:
                    for cell in row:
                        ws2[cell.coordinate].value = cell.value
    wb.save('merged.xlsx')


if __name__ == '__main__':
    merge()
